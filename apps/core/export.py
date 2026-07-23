import io
from datetime import date
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def build_workbook(sheets, filename):
    """sheets: list of {title, headers, rows} where rows is list of [label, *values]"""
    wb = Workbook()
    wb.remove(wb.active)

    for idx, sheet_data in enumerate(sheets):
        title = sheet_data["title"][:31]  # Excel limit
        # Sanitize invalid chars for Excel sheet names: \ / * ? : [ ]
        for ch in ['\\', '/', '*', '?', ':', '[', ']']:
            title = title.replace(ch, '-')
        ws = wb.create_sheet(title=title, index=idx)

        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers) + 1)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = TITLE_FONT

        # Headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx + 1, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        # Indicator label header
        label_cell = ws.cell(row=2, column=1, value="Indicador")
        label_cell.font = HEADER_FONT
        label_cell.fill = HEADER_FILL
        label_cell.border = THIN_BORDER

        # Data rows
        for row_idx, row in enumerate(rows, start=3):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = CELL_FONT
                cell.border = THIN_BORDER
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = Alignment(horizontal="center")

        # Column widths
        ws.column_dimensions["A"].width = 42
        for col_idx in range(2, len(headers) + 2):
            col_letter = chr(64 + col_idx) if col_idx <= 26 else "A" + chr(64 + col_idx - 26)
            ws.column_dimensions[col_letter].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


class ExcelExportMixin:
    export_filename = "dados.xlsx"

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") == "xlsx" and getattr(self, "is_admin", lambda: False)():
            return self.export_excel()
        return super().get(request, *args, **kwargs)
